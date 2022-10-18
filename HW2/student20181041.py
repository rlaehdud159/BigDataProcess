#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

#total 계산
row_id = 1
for row in ws:
  if row_id != 1:
    sum_v = ws.cell(row = row_id, column = 3).value * 0.3
    sum_v += ws.cell(row = row_id, column = 4).value * 0.35
    sum_v += ws.cell(row = row_id, column = 5).value * 0.34
    sum_v += ws.cell(row = row_id, column = 6).value
    ws.cell(row = row_id, column = 7).value = sum_v

  row_id += 1

#total 정렬
sum_list = []
sum_sort_list = []
for i in range(2, ws.max_row + 1):
  sum_list.append(ws.cell(row = i, column = 7).value)
sum_sort_list = sorted(sum_list, reverse = True)

#총 수강생 수
count = 0
for x in range(2, ws.max_row + 1):
  count += 1

#grade 계산
a_count = count * 0.3
a = sum_sort_list[a_count - 1]
aplus_count = a_count * 0.5
aplus = sum_sort_list[aplus_count - 1]

b_count = count * 0.7
b = sum_sort_list[b_count - 1]
bplus_count = a_count + (b_count - a_count) * 0.5
bplus = sum_sort_list[bplus_count - 1]

cplus_count = b_count + (count * 0.3 * 0.5)
cplus = sum_sort[cplus_count - 1]

row_id = 1
for row in ws:
  if row_id != 1:
    sum_v = ws.cell(row = row_id, column = 7).value
    grade = 'C0'

    if(sum_v >= aplus):
      grade = 'A+'
    elif(sum_v >= a):
      grade = 'A'
    elif(sum_v >= bplus)
      grade = 'B+'
    elif(sum_v >= b)
      grade = 'B'
    elif(sum_v >= cplus)
      grade = 'C+'
   
    ws.cell(row = row_id, column = 8).value = grade
  row_id += 1
 
wb.save("student.xlsx")
